import time
import os
from selenium import webdriver
# For Email function for mapcode
# import smtplib
# for backup email
# from email.mime.base import MIMEBase
# from email.mime.multipart import MIMEMultipart
# from email.mime.text import MIMEText
# from email import encoders
import tkinter.messagebox
from Customer import Customer
from openpyxl import load_workbook
import win32com.client
from selenium.common.exceptions import ElementNotInteractableException

deps = f"C:/Users/{os.getlogin()}/OneDrive - Hexagon/Deps_xSave/"


class Email:

    def __init__(self):
        print('Email Class created')

    # @staticmethod
    # def send_zipped_backup_email(email_recipient):
    #     wb = load_workbook(filename=os.path.join(deps, 'Details/DsName.xlsx'))
    #     ws = wb.active
    #     print('values below are from workbook load')
    #     print(ws.cell(row=1, column=2).value)
    #     dsd = ws.cell(row=1, column=2).value
    #     print(ws.cell(row=2, column=2).value)
    #     zippo = ws.cell(row=2, column=2).value
    #     wb.close()
    #
    #     # em_mac = Customer.email_mac_info()
    #     # em_mac = Customer.Customer.email_mac_info()
    #     # em_mac_type = em_mac[1]
    #     # em_mac_size = em_mac[3]
    #     # em_mac_serial = em_mac[4]
    #     # em_customer = Customer.customer_info()
    #
    #     my_email = 'lee.maloney@hexagon.com'
    #     # email_recipient = "ServiceAdmin.uk@hexagon.com"
    #     # email_recipient = emailRecipientAdd
    #     subject = "CMM Backup files"
    #
    #     msg = MIMEMultipart()
    #     msg["From"] = my_email
    #     msg["To"] = email_recipient
    #     msg["Subject"] = subject
    #
    #     body = 'Please see attached machine backup files for ' + Customer.get_customer_machine_model() + " " + \
    #            Customer.get_customer_machine_size() + ', Serial N.o:' + Customer.get_customer_machine_serial() + \
    #            " at " + Customer.get_customer_name() + ', ' + Customer.get_customer_address() + "." + "\n\n" \
    #            + "Regards" + "\nLee"
    #
    #     msg.attach(MIMEText(body, 'plain'))
    #
    #     email_attach = dsd + zippo
    #     attachment = open(email_attach, 'rb')
    #     print("level2")
    #     part = MIMEBase('application', 'octet-stream')
    #     part.set_payload(attachment.read())
    #     encoders.encode_base64(part)
    #     part.add_header('Content-Disposition', "attachment; filename=" + zippo)
    #     msg.attach(part)
    #
    #     text = msg.as_string()
    #     server = smtplib.SMTP("smtp.office365.com", port=587)
    #     server.starttls()
    #     server.login(my_email, 'Lancaster9')
    #
    #     send_email = tkinter.messagebox.askyesno("Email", f'Backup for: \n{Customer.get_customer_machine_model()}, '
    #                                                       f'{Customer.get_customer_machine_size()}, Serial N.o:'
    #                                                       f'{Customer.get_customer_machine_serial()} at '
    #                                                       f'\n{Customer.get_customer_name()}, '
    #                                                       f'\n{Customer.get_customer_address()} \nSend Email?')
    #     if send_email:
    #         server.sendmail(my_email, email_recipient, text)
    #         server.quit()
    #         print("server quit")

    @classmethod
    def send_zipped_backup_email(cls, email_recipient):
        wb = load_workbook(filename=os.path.join(deps, 'Details/DsName.xlsx'))
        ws = wb.active
        print('values below are from workbook load')
        print(ws.cell(row=1, column=2).value)
        datasave_dir = ws.cell(row=1, column=2).value
        print(ws.cell(row=2, column=2).value)
        zip_file_name = ws.cell(row=2, column=2).value
        wb.close()

        body = 'Please see attached machine backup files for ' + Customer.get_customer_machine_model() + " " + \
               Customer.get_customer_machine_size() + ', Serial N.o: ' + Customer.get_customer_machine_serial() + \
               " at " + Customer.get_customer_name() + ', ' + Customer.get_customer_address() + "." + "\n\n" \
               + "Regards" + "\n" + os.getlogin().split(".")[0]

        email_attach = datasave_dir + zip_file_name
        send_email = tkinter.messagebox.askyesno("Email", f'Backup for: \n{Customer.get_customer_machine_model()}, '
                                                          f'{Customer.get_customer_machine_size()}, Serial N.o: '
                                                          f'{Customer.get_customer_machine_serial()}\nAt:'
                                                          f'\n{Customer.get_customer_name()}, '
                                                          f'{Customer.get_customer_address()} \n\nSend Email?')

        if send_email:
            try:
                os.system("TASKKILL /F /IM Outlook.exe")
            except Exception as ex:
                print(f'exception is {ex}')
            time.sleep(1)
            # outlook = win32com.client.gencache.EnsureDispatch('Outlook.Application')
            outlook = win32com.client.Dispatch("Outlook.Application")
            message = outlook.CreateItem(0)
            message.Display()
            message.To = email_recipient
            message.Subject = f'CMM Backup files: {Customer.get_customer_name()}'
            message.Body = body
            message.Attachments.Add(Source=str(email_attach))
            # message.Display()
            try:
                message.Send()
                os.startfile("outlook")
            except Exception as ex:
                print(f'exception is {ex}')
            print("end of function")

    @classmethod
    def get_and_email_fdc_code(cls):
        get_code = tkinter.messagebox.askyesno("Map code", "Get map code from controller")
        if get_code:
            driver = webdriver.Chrome(executable_path=os.path.join(deps, 'chromedriver_win32/chromedriver.exe'))
            driver.implicitly_wait(30)
            driver.get("http://100.0.0.1")
            driver.maximize_window()
            # Log-in
            time.sleep(3.5)
            try:
                driver.find_element_by_id("login_PassText").send_keys("hexmet")
                time.sleep(1.0)
            except ElementNotInteractableException:
                driver.find_element_by_id("errorButton1").click()
                time.sleep(1.5)
                driver.find_element_by_id("login_PassText").send_keys("hexmet")
                time.sleep(1.0)
            driver.find_element_by_id("login_BtnLogin").click()
            time.sleep(0.5)
            # click on tools menu
            driver.find_element_by_id("ui-accordion-mainMenu-header-2").click()
            time.sleep(0.5)
            try:
                driver.find_element_by_link_text('Map Management').click()
                # driver.execute_script('mainMenu_OnClick("MapManagement.asp")')
                # time.sleep(2)
            except Exception as ex:
                print(f'Exception is {ex}')
                # driver.execute_script('mainMenu_OnClick("MapManagement.html")')
            time.sleep(0.3)
            driver.find_element_by_id("mapManagement_GetButton").click()
            time.sleep(0.3)
            driver.find_element_by_id("mapManagement_CIDText").click()
            mapcode = driver.find_element_by_id("mapManagement_CIDText").get_attribute("value")

            try:
                os.system("TASKKILL /F /IM Outlook.exe")
                time.sleep(0.5)
            except Exception as ex:
                print(f'exception is {ex}')
            # time.sleep(1)

            # try:
            # outlook = win32com.client.GetObject('Outlook.Application')
            # except:
            #     print("EX")
            #     outlook = win32com.client.Dispatch('Outlook.Application')

            # outlook = win32com.client.gencache.EnsureDispatch('Outlook.Application')
            # outlook = win32com.client.GetActiveObject('Outlook.Application')
            outlook = win32com.client.Dispatch("Outlook.Application")
            message = outlook.CreateItem(0)
            message.Display()
            message.To = 'fdcbackup2013@gmail.com'
            message.Subject = 'REQUEST_KEY'
            message.Body = f'User Code:2ffa2c1131cd4b1d9a30c6c2a7b3a90da1d60b9c6383430f976f91a4292231c3\n' \
                           f'Controller ID:{mapcode}\n'
            # message.Display()
            try:
                message.Send()
                time.sleep(0.5)
                os.startfile("outlook")
            except Exception as ex:
                print(f'exception is {ex}')
            print("end of function")

            # conn = smtplib.SMTP("smtp.office365.com", port=587)
            # conn.ehlo()
            # conn.starttls()
            # conn.login('lee.maloney@hexagon.com', 'Lancaster9')
            # conn.sendmail('lee.maloney@hexagon.com', 'lee.maloney@hexagon.com',
            #               'Subject:REQUEST_KEY\n\n User '
            #               'Code:2ffa2c1131cd4b1d9a30c6c2a7b3a90da1d60b9c6383430f976f91a4292231c3 '
            #               + "\n" + "Controller ID: " + mapcode + "\n ")
            # conn.quit()
