# datasave commit

import PyPDF2
import tkinter.messagebox
from tkinter import simpledialog
import datetime
import openpyxl
import shutil
import os
import win32print
import glob

import constants

deps = constants.dependencies_location
# deps = f"C:/Users/{os.getlogin()}/OneDrive - Hexagon/Deps_xSave/"


class Datasave:

    def __init__(self):
        print('Datasave Class Created')

    # def gen_datasave_directory_name(c):
    #     isinstance(c, str)
    #     datasavedir = ('C:\\Users\\lee.maloney\\OneDrive - Hexagon\\CalCerts and Health checks\\2021\\'
    #                    + str(c) + "_" + cal_config.model + "_" + str(date_time) + "\\")
    #     cal_config.set_datasave_path_name(datasavedir)
    #     return datasavedir

    @staticmethod
    def cleanup_datasave_autotune(dir_name):
        autotune_dir = os.listdir(dir_name)
        autotune_remove_list = [".exe", ".dll", ".config", ".zip", ".prc", ".ht", ".tab", ".bat"]
        for item in autotune_dir:
            for _ in autotune_remove_list:
                if item.lower().endswith(_):
                    os.remove(os.path.join(dir_name, item))
        for item in autotune_dir:
            if item.lower().endswith('data') or item.lower().endswith('v05.5'):
                shutil.rmtree(os.path.join(dir_name, item))

    @staticmethod
    def move_scan_results_to_datasave(cal_config):
        # todo - Check this on next scanning machine
        if cal_config.get_controller_type() == 'DC':
            list_of_files = glob.glob('C:/DeaReport/dataexport/*.pdf')  # * means all if need specific format then *.csv
            latest_file_path = max(list_of_files, key=os.path.getctime)
            scan_file_name = os.path.basename(latest_file_path)
            shutil.copyfile(latest_file_path,
                            cal_config.datasave_path + 'Certificates and Documents\\' + str(scan_file_name))
        if cal_config.get_controller_type() == 'CC' or cal_config.get_controller_type() == "CC_Hyper":
            list_of_files = glob.glob('"C:/SIQFILES"/*.pdf')  # * means all if need specific format then *.csv
            latest_file_path = max(list_of_files, key=os.path.getctime)
            scan_file_name = os.path.basename(latest_file_path)
            shutil.copyfile(latest_file_path,
                            cal_config.datasave_path + 'Certificates and Documents\\' + str(scan_file_name))

    @staticmethod
    def move_mdb_to_datasave(cal_config):
        list_of_files = glob.glob('C:/DeaReport/dataexport/*.mdb')  # * means all if need specific format then *.csv
        latest_file_path = max(list_of_files, key=os.path.getctime)
        mdb_file_name = os.path.basename(latest_file_path)
        print(f"latest mdb file full path: {latest_file_path}, base file: {mdb_file_name}")
        shutil.copyfile(latest_file_path, str(cal_config.get_datasave_path_name()) +
                        'Certificates and Documents\\' + mdb_file_name)

    @staticmethod
    def generate_cal_label(cal_config):
        atreport_location = (os.path.join(cal_config.get_datasave_path_name(), "Certificates and "
                                                                               "Documents/ATReport50.pdf"))
        print(atreport_location)
        if os.path.isfile(atreport_location):
            standard, dash3, dash4, = '', '', ''
            front_iso_part = 'ISO 10360-2'
            std_2001 = tkinter.messagebox.askyesno("Iso Standard", "Is this machine calibrated to ISO10360-2:2001?")
            if std_2001:
                standard = ':2001'
            else:
                std_2009 = tkinter.messagebox.askyesno("Iso Standard",
                                                       "Is this machine calibrated to ISO10360-2:2009?")
                if std_2009:
                    standard = ':2009'
                else:
                    std_cmma = tkinter.messagebox.askyesno("Standard", "Is this machine calibrated to CMMA?")
                    if std_cmma:
                        standard = "CMMA"
                        front_iso_part = ''
                    else:
                        std_vdi_vde = tkinter.messagebox.askyesno("Standard", "Is this machine calibrated to VDI?")
                        if std_vdi_vde:
                            standard = "VDI/VDE 2617"
                            front_iso_part = ''
            dash_3_performed = tkinter.messagebox.askyesno("Rotary Table Test", "Was a -3 test performed and passed?")
            if dash_3_performed:
                dash3 = '-3'
            dash_4_performed = tkinter.messagebox.askyesno("Scanning Test", "Was a -4 test performed and passed?")
            if dash_4_performed:
                dash4 = '-4'

            iso = front_iso_part + dash3 + dash4 + standard

            datasavedir = cal_config.get_datasave_path_name()
            # ----------------- Extract cal cert number
            with open(datasavedir + 'Certificates and Documents\\' + "ATReport50.pdf", 'rb') as pdfFileObj:
                # print(pdfFileObj)
                pdfreader = PyPDF2.PdfFileReader(pdfFileObj)
                pageobj = pdfreader.getPage(0)
                page = pageobj.extractText()
                line = str(page)
                # print(line.encode("utf-8"))
                p1 = line.split(' ')
                certno = p1[8]
                print("Cert number for cal label is " + str(certno))
            # ----------------- Get machine spec
            with open('C:\\pcdmisw\\pcddata\\isotol.dat', 'r') as f:
                firstline = f.readline()
                secondline = f.readline()
                thirdline = f.readline()
                fourthline = f.readline()
                firstline.split(' ')
                b = secondline.split(' ')
                c = thirdline.split(' ')
                d = fourthline.split(' ')
                firstterm = float(b[1]) * 1000
                print(str(float(c[1][0:8])))
                secondterm = 1000 / ((float(c[1][0:8])) * 1000)
                secondterm = round(secondterm, 2)
                print("sec term is " + str(secondterm))
                probeerror = float(d[1]) * 1000
                spec = "+/-(" + str(firstterm) + " + L/" + str(secondterm) + "), MPE=" + str(probeerror)
                print(spec)
            # ----------------- Get Serial number
            with open('C:\\pcdmisw\\pcddata\\isomac.dat', 'r') as f:
                f.readline()
                f.readline()
                f.readline()
                f.readline()
                fifth = f.readline()
                serialno = fifth
                print("serial no is " + serialno)
            is_rlc_machine = tkinter.messagebox.askyesno("RLC", "Is This an RLC machine?")
            if is_rlc_machine:
                z_number = simpledialog.askstring("Z Number", "Enter Z number")
                plant_number = simpledialog.askstring("Plant Number", "Enter plant number")
                serialno = f"{serialno}/{z_number}/{plant_number}"

            # ----------------- open master insert values and save as CalLabel
            wb = openpyxl.load_workbook(filename=os.path.join(deps, 'Cal_forms/masterCalLabel.xlsx'))
            ws = wb.active
            ws['e3'] = certno
            ws['e4'] = serialno
            dtcal = datetime.datetime.now()
            ws['e5'] = dtcal.strftime("%d" + "/" + "%m" + "/" + "%Y")
            nxtyr = datetime.datetime.now().strftime("%Y")
            nxtyr = int(nxtyr) + 1
            print(nxtyr)
            ws['e6'] = datetime.datetime.now().strftime("%d" + "/" + "%m") + "/" + str(nxtyr)
            ws['e7'] = iso
            ws['e8'] = 'Yes'
            wb.save(datasavedir + 'Certificates and Documents\\' +
                    'KP3-STA-FM-073 - CMM Calibration Label - master non ukas - issue 01 07 12 20.xlsx')
            print(f'saving to {datasavedir} /Certificates and Documents\\callabel')
            wb.close()
        else:
            tkinter.messagebox.showerror("xSave helper", "Ensure Atreport50.pdf has been generated and is present in "
                                                         "Certificates and documents.")

    @staticmethod
    def convert_files_to_pdf(cal_config):
        win32print.SetDefaultPrinter("CutePDF Writer")
        dsd = os.path.join(cal_config.get_datasave_path_name(), "Certificates and Documents/")
        # dsd = cal_config.get_datasave_path_name()
        os.chdir(dsd)
        try:
            os.startfile(dsd + "KP3-STA-FM-073 - CMM Calibration Label - master non ukas - issue 01 07 12 20.xlsx",
                         "print")
            os.startfile(dsd + "KP3-STA-FM-094 - Service Check Sheet  Issue 01  02 09 19.docx", "print")
        except FileNotFoundError:
            print('One of files not found')
            print(dsd)

    @staticmethod
    def create_datasave_zipfile(cal_config):
        # shutil.make_archive makes a zip file and puts it in the same directory as the file to be zipped
        azip = shutil.make_archive(cal_config.datasave_path, 'zip', cal_config.datasave_path)
        print("azip is : " + azip)
        list_dsd = os.listdir(cal_config.datasave_path)
        for item in list_dsd:
            if item.lower().endswith(".zip"):
                os.remove(cal_config.datasave_path + item)
        shutil.move(azip, cal_config.datasave_path)
        list_dsd = os.listdir(cal_config.datasave_path)
        for zippo in list_dsd:
            if zippo.lower().endswith(".zip"):
                print("dsd + zippo equals " + str(cal_config.datasave_path) + str(zippo))
                cal_config.put_detail_into_workbook(location='b2', detail=str(zippo))

    @staticmethod
    def copy_autotune_files(cal_config):
        autotune_list = []
        # x = cal_config.get_autotune_c_drive_folder_name()
        if cal_config.get_autotune_c_drive_folder_name() == 'AutotuneDC':
            autotune_list = ['adjtest1.txt', 'adjtest2.txt', 'adjtest3.txt', 'autadj.log', 'paramCommon.xml',
                             'paramHw.xml', 'paramSys.xml', 'paramW.xml', 'paramX.xml', 'paramY.xml', 'paramZ.xml',
                             'pcdparams.txt',  'prbparams.txt', 'sensor.asc', 'serv.stp', 'Temploadok.dat']
        if cal_config.get_autotune_c_drive_folder_name() == 'Autotune':
            autotune_list = ['adjtest1.txt', 'adjtest2.txt', 'adjtest3.txt', 'autadj.log', 'CONSTANT.ASC',
                             'CREACOS.ASC', 'DEACFG.ASC', 'E2POT.DAT', 'E2POTSCL.DAT', 'pcdparams.txt', 'sensor.asc',
                             'serv.stp']
        for file in os.listdir(cal_config.get_autotune_dir()):
            if file in autotune_list or file.lower().endswith('adj'):
                source = os.path.join(cal_config.get_autotune_dir(), file)
                destination = cal_config.get_datasave_path_name() + os.path.join(
                    cal_config.get_autotune_c_drive_folder_name(), file)
                shutil.copyfile(source, destination)
