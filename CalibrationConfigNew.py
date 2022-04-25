from openpyxl import load_workbook
import os
import shutil
import tkinter.messagebox

import constants

deps = constants.dependencies_location


class CalConfig2:
    model: str = ''
    serial_number: str = ''
    controller: str = ''
    zip_filename: str = ''
    datasave_path: str = ''
    backup_name: str = ''
    autotune_dir: str = ''
    autotune_fold_name: str = ''

    # read last saved config and pull in the data
    # run these two in main
    # cls.load_workbook_details_to_calconfig_object()
    # set_autotune_data()

    @classmethod
    def about(cls):
        return print(f'Cal config is a {cls.model}, with a {cls.controller}, s.n:{cls.serial_number}, \n datasave '
                     f'located at {cls.datasave_path}, \nZip filename is {cls.zip_filename}')

    @classmethod
    def get_model(cls) -> str:
        return cls.model

    @classmethod
    def set_serial_number(cls, serial_no):
        cls.serial_number = serial_no
        cls.put_detail_into_workbook(location='b8', detail=serial_no)

    @classmethod
    def get_serial_number(cls):
        return cls.serial_number

    @classmethod
    def set_datasave_path_name(cls, path_name, backup_name):
        cls.datasave_path = path_name
        cls.backup_name = backup_name
        print(f'datasave names are {cls.datasave_path}, {cls.backup_name}')
        cls.put_detail_into_workbook(location='b1', detail=str(path_name))
        cls.put_detail_into_workbook(location='b3', detail=str(backup_name))

    @classmethod
    def get_datasave_path_name(cls):
        return cls.datasave_path

    @classmethod
    def get_backup_name(cls):
        return cls.backup_name

    @classmethod
    def set_autotune_data(cls):
        if cls.controller == 'CC Controller' or cls.controller == "CC Controller_Hyper":
            cls.autotune_dir = 'c:/Autotune'
            cls.autotune_fold_name = 'Autotune'
        if cls.controller == 'DC Controller':
            if "AutotuneDC" in os.listdir("C:/"):
                cls.autotune_dir = 'c:/AutotuneDC'
                cls.autotune_fold_name = 'AutotuneDC'
            if "autotunedc" in os.listdir("C:/"):
                cls.autotune_dir = 'c:/autotunedc'
                cls.autotune_fold_name = 'autotunedc'

    @classmethod
    def get_autotune_dir(cls):
        cls.set_autotune_data()
        return cls.autotune_dir

    @classmethod
    def get_autotune_c_drive_folder_name(cls):
        cls.set_autotune_data()
        return cls.autotune_fold_name

    @classmethod
    def get_controller_type(cls):
        if cls.controller == 'CC Controller':
            return 'CC'
        elif cls.controller == "CC Controller_Hyper":
            return "CC_Hyper"
        elif cls.controller == 'DC Controller':
            return 'DC'

    @classmethod
    def load_workbook_details_to_calconfig_object(cls):
        wb = load_workbook(filename=os.path.join(deps, 'Details/DsName.xlsx'))
        ws = wb.active
        cls.datasave_path = ws.cell(row=1, column=2).value
        cls.backup_name = ws.cell(row=3, column=2).value
        cls.zip_filename = ws.cell(row=2, column=2).value
        cls.model = ws.cell(row=5, column=2).value
        cls.controller = ws.cell(row=6, column=2).value
        cls.serial_number = ws.cell(row=8, column=2).value
        cls.set_autotune_data()
        wb.close()

    @classmethod
    def put_detail_into_workbook(cls, location, detail):
        wb = load_workbook(filename=os.path.join(deps, 'Details/DsName.xlsx'))
        ws = wb.active
        ws[location] = str(detail)
        wb.save(os.path.join(deps, 'Details/DsName.xlsx'))
        wb.close()
        cls.load_workbook_details_to_calconfig_object()

    @staticmethod
    def clear_workbook_details():
        wb = load_workbook(filename=os.path.join(deps, 'Details/DsName.xlsx'))
        ws = wb.active
        ws['b5'], ws['b6'], ws['b7'], ws['b8'] = " ", " ", " ", " "
        wb.save(os.path.join(deps, 'Details/DsName.xlsx'))
        wb.close()

    @staticmethod
    # Clear last calibration files from Laptop. Runs through each folder targeting specific files and extensions.
    def clear_last_cal_files():
        dir_list = ["c:/tutor", "c:/AutotuneDC", "c:/autotunedc", "c:/Autotune", "c:/servicedea/cmmdata/arm1",
                    "c:/Program Files/Thermal_ocx", "c:/PCDMISW/PCDDATA", "c:/DeaReport/dataexport", "c:/PCDMISW/pcdpp"]
        for dir_name in dir_list:
            if dir_name == "c:/Autotune":
                autotune_ext_list = [".adj", ".stp"]
                for ext in autotune_ext_list:
                    directory_name = os.listdir(dir_name)
                    for item in directory_name:
                        if item.lower().endswith(ext) or item.lower() in constants.autotune_ext_delete_list:
                            os.remove(os.path.join(dir_name, item))
            if dir_name == "c:/autotunedc" or dir_name == "c:/AutotuneDC":
                autotunedc_ext_list = [".stp", ".adj", "ctr.asc"]
                for ext in autotunedc_ext_list:
                    directory_name = os.listdir(dir_name)
                    for item in directory_name:
                        if item.lower().endswith(ext) or item.lower() in constants.autotunedc_ext_delete_list:
                            os.remove(os.path.join(dir_name, item))
            _directory = os.listdir(dir_name)
            for item in _directory:
                if dir_name == "c:/Program Files/Thermal_ocx":
                    if item.lower().endswith(".stp"):
                        os.remove(os.path.join(dir_name, item))
                if dir_name == "c:/servicedea/cmmdata/arm1":
                    if item.endswith(".dat") or item.lower().endswith(".asc") or item.lower().endswith("file.txt"):
                        os.remove(os.path.join(dir_name, item))
                if dir_name == "c:/tutor":
                    if item.lower().endswith(".dat") or item.lower().endswith(".asc"):
                        os.remove(os.path.join(dir_name, item))
                if dir_name == "c:/PCDMISW/PCDDATA":
                    if item.lower() in constants.pcddata_delete_list or item.lower().endswith(".qiq"):
                        os.remove(os.path.join(dir_name, item))
                if dir_name == "c:/DeaReport/dataexport":
                    if item.lower().endswith(".pdf") or item.lower().endswith(".hrp") or item.lower().endswith(".mdb"):
                        os.remove(os.path.join(dir_name, item))
                if dir_name == "c:/PCDMISW/pcdpp":
                    if item.endswith("testlm.PRG"):
                        os.remove(os.path.join(dir_name, item))
                if dir_name == "c\\SIQFILES":
                    if item.lower().endswith("siq") or item.lower().endswith(".pdf"):
                        os.remove(os.path.join(dir_name, item))

    @staticmethod
    def disable_uncertainty():
        with open('C:/Program Files (x86)/DeaReport/ATReport.ini', 'r') as file:
            uncertainty_ini_file = file.readlines()
        uncertainty_ini_file[2] = 'DisabilitaIncertezza=1\n'
        with open('C:/Program Files (x86)/DeaReport/ATReport.ini', 'w') as file:
            file.writelines(uncertainty_ini_file)

    @staticmethod
    def enable_uncertainty():
        with open('C:/Program Files (x86)/DeaReport/ATReport.ini', 'r') as file:
            uncertainty_ini_file = file.readlines()
        uncertainty_ini_file[2] = 'DisabilitaIncertezza=0\n'
        with open('C:/Program Files (x86)/DeaReport/ATReport.ini', 'w') as file:
            file.writelines(uncertainty_ini_file)

    # remove redundant if item.lower() == "serv.stp": and use a try catch with a file not found exception.
    #  Refactor with propper names for sizes.

    @classmethod
    def copy_serv_file(cls):
        atdir = os.listdir(cls.get_autotune_dir())
        print(atdir)
        for item in atdir:
            if item.lower() == "serv.stp":
                with open(os.path.join(cls.get_autotune_dir(), item)) as f:
                    for i in range(1, 8):
                        sizes = f.readline()
                    for i in range(1, 9):
                        f.readline()
                    line_sixteen = f.readline()
                    line_sixteen = line_sixteen.strip('\n')
                    line_seventeen = f.readline()
                    line_seventeen = line_seventeen.strip('\n')
                    sizes = str(sizes).strip('\n')
                    sizes = sizes.replace('.0', '')
                    sizes = sizes.split(' ')
                copy_serv = tkinter.messagebox.askyesno("Copy Serv.STP file?", f'Machine data from uploaded serv file:'
                                                                               f'\n\nLyy: {sizes[2]}\nLxx: {sizes[1]}'
                                                                               f'\nLzz:  {sizes[3]}'
                                                                               f'\n\nGranite: {line_sixteen}'
                                                                               f'\nBridge:  {line_seventeen}\n\n'
                                                                               f'Do you want to copy'
                                                                               f' serv file to Thermal_OCX?')
                if copy_serv:
                    shutil.copyfile(os.path.join(cls.get_autotune_dir(), item), "C:\\Program Files\\Thermal_ocx\\"
                                                                                "Serv1.Stp")
                    print("Serv file copied dc")
                    os.startfile("C:\\Program Files\\Thermal_ocx")
        return sizes

    @staticmethod
    def delete_probes():
        for parent, subdir, files in os.walk('C:/Users/Public/Documents/Hexagon/PC-DMIS'):
            for file in files:
                if file.endswith(".PRB") or file.endswith(".Results"):
                    os.remove(os.path.join(parent, file))
